# -*- coding: utf-8 -*-
__author__ = "ruijing.li"

import pandas as pd
import pyreadstat
import datacompy
import operator
import logging
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Border, Side, Alignment
from common.excel_format import content_style, content_heading_style, content_value_style
from tkinter import filedialog
from openpyxl.utils import get_column_letter
from collections import defaultdict
from common.readConfig import MyConfig

def read_sas(path, data):
    df, meta = pyreadstat.read_sas7bdat(f"{path}{data}.sas7bdat")
    return df, meta

def read_excel(path, data):
    df = pd.read_excel(f"{path}{data}.xlsx")
    return df

def read_content(PATH_INPUT_Config) -> dict:
    oid_list = MyConfig(PATH_INPUT_Config).content_formOID
    form_name = MyConfig(PATH_INPUT_Config).content_formName
    form_dict = dict(zip(oid_list, form_name))
    return form_dict

def read_content_excel(PATH_INPUT) -> dict:
    content = pd.read_excel(PATH_INPUT, sheet_name='Sheet1')
    oid_list = content['OID'].tolist()
    form_name = content['FORMNAME'].tolist()
    form_dict = dict(zip(oid_list, form_name))
    return form_dict

def read_visitorder_excel(PATH_INPUT) -> pd.DataFrame:
    content = pd.read_excel(PATH_INPUT, sheet_name='Sheet1')
    try:
        visitoid_list = content['VISTOID'].tolist()
        visitno = content['VISITNO'].tolist()
        visit_order = pd.DataFrame({'visitno': visitno, 'vistoid': visitoid_list})
        logging.warning("已读取visit顺序信息...")
    except Exception:
        visit_order = None
        logging.warning("未读取到visit顺序信息...")
    return visit_order

def create_content(ws,FORMS):
    ws.merge_cells("A1:C1")
    ws["A1"] = "Contents"
    ws["A2"] = "Sequence"
    ws["B2"] = "OID"
    ws["C2"] = "Forms"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50

    for cell in ws[1]:  # content
        cell.style = content_style

    for cell in ws[2]:
        cell.style = content_heading_style

    row = 3
    for key, values in FORMS.items():
        ws[f"A{row}"] = row - 2

        ws[f"B{row}"] = key
        ws[f"C{row}"] = values
        ws[f"A{row}"].style = content_value_style
        ws[f"B{row}"].style = content_value_style
        ws[f"C{row}"].style = content_value_style
        row += 1

    set_content_hyperlink(worksheet=ws, start_row=3, end_row=row, link_col="B")


def set_content_hyperlink(worksheet, start_row: int, end_row: int, link_col: str) -> None:
    for i in range(start_row, end_row):
        cell_tar = worksheet["{0}{1}".format(link_col, i)]
        cell_tar.hyperlink = "#{0}!A1".format(cell_tar.value)
        cell_tar.style = "Hyperlink"

def set_worksheet_format(ws,outputFLAG):
        cell_tar = ws.cell(row = 1,column = ws.max_column+1,value="BackToContents")
        cell_tar.hyperlink = "#{0}!A1".format("Contents")
        cell_tar.style = "Hyperlink"

        yellowFill = PatternFill(start_color='FFFF00',
                              end_color='FFFF00',
                              fill_type='solid')

        cell_tar2 = ws.cell(row = 1,column = ws.max_column-1)
        # cell_tar2 = ws.cell(row = 1,column = ws.max_column)
        cell_tar2.fill  = yellowFill

        ws.row_dimensions[1].height = 50

        for col in ws.columns:
            index = list(ws.columns).index(col)
            letter = get_column_letter(index + 1)
            ws.column_dimensions[letter].width = 14
            ws[f"{letter}1"].alignment = Alignment(wrapText=True)

        ws.auto_filter.ref = ws.dimensions

        if outputFLAG == "批注":
            for row in ws.iter_rows():
                for cell in row:
                    if operator.contains(str(cell.value), "->"):
                        if str(cell.value).split("->")[0].strip() == "nan" or str(cell.value).split("->")[0].strip() == "NaT":
                            comment = Comment("", "system")
                        else:
                            comment = Comment(str(cell.value).split("->")[0].strip(), "system")
                        cell.comment = comment
                        cell.value = str(cell.value).split("->")[1].strip()


def get_compare_variables_Common(data: pd.DataFrame, not_compare_variables: list,system: list):
    not_compare_variables_1 = [x.upper() for x in not_compare_variables]
    filtered_not_compare_variables = [var for var in not_compare_variables_1 if var in data]
    not_compare = system + filtered_not_compare_variables
    not_compare_1 = [x.upper() for x in not_compare]
    compare_variables = []

    def not_in_list(variable: str, variables: list):
        FLAG = True
        for i in variables:
            if i == variable:
                FLAG = False
                break
        return FLAG

    for column in data.columns:
        if not_in_list(column, not_compare_1):
            compare_variables.append(column)
    return compare_variables

def get_not_compare_data(data:pd.DataFrame,variable:list) -> pd.DataFrame:
    TS = data[variable]
    seq2 = list(map(lambda x: x.lower(), variable))
    seq1 = variable
    LABEL_LIST = dict(zip(seq1, seq2))
    TS = TS.rename(columns=LABEL_LIST)
    return TS

def compare_data_common(OldData: pd.DataFrame, NewData: pd.DataFrame, Variables: list,
                        not_compare_variables: list, system_variables: list,
                        system_not_compare: list,KEEP_DELETE_DATA:bool,COLOUR:bool) -> pd.DataFrame:
    not_compare_variables_2 = [x.upper() for x in not_compare_variables]
    filtered_not_compare_variables = [var for var in not_compare_variables_2 if var in NewData]

    compare_variables = [*system_variables, *Variables]
    not_compare_variables_list_1 = [*system_variables, *system_not_compare, *filtered_not_compare_variables]
    not_compare_variables_list = [x.upper() for x in not_compare_variables_list_1]

    OldDataSet = OldData.fillna(' ').applymap(str)
    NewDataSet = NewData.fillna(' ').applymap(str)

    not_compare_dataframe_Old = get_not_compare_data(OldDataSet, not_compare_variables_list)
    not_compare_dataframe_New = get_not_compare_data(NewDataSet, not_compare_variables_list)

    OldDataSet = OldDataSet.drop(columns=list(set(OldDataSet.columns).difference(set(compare_variables))))
    NewDataSet = NewDataSet.drop(columns=list(set(NewDataSet.columns).difference(set(compare_variables))))

    compare_AddOrDelete = datacompy.Compare(OldDataSet, NewDataSet, join_columns=system_variables)

    data_deleted = compare_AddOrDelete.df1_unq_rows
    data_add = compare_AddOrDelete.df2_unq_rows

    data_compare_Change = datacompy.Compare(OldDataSet, NewDataSet, join_columns=compare_variables)

    data_new = data_compare_Change.df2_unq_rows
    data_old = data_compare_Change.df1_unq_rows

    data_unchange = data_compare_Change.intersect_rows
    data_unchange = data_unchange.drop(["_merge"], axis=1)

    data_new = pd.concat([data_new, data_add], ignore_index=True).drop_duplicates(keep=False)
    data_old = pd.concat([data_old, data_deleted], ignore_index=True).drop_duplicates(keep=False)

    # data_new = pd.concat([data_new, data_add], ignore_index=True)
    # data_old = pd.concat([data_old, data_deleted], ignore_index=True)

    merge_variables = list(map(lambda x: x.lower(), system_variables))

    data_concat = pd.merge(data_new, data_old, how="left", on=merge_variables)

    data_change = pd.DataFrame()

    for (colname, colval) in data_concat.items():
        if len(data_concat) == 0:
            break

        elif colname.endswith("_x"):
            prefix = colname.replace("_x", "")
            New = prefix + "_x"
            Old = prefix + "_y"
            if COLOUR:
                data_change[prefix] = data_concat.apply(
                    lambda x: str(x[Old]) + " -> " + str(x[New]) if str(x[New]) != str(x[Old]) else str(x[New]),
                    axis=1)
            else:
                data_change[prefix] = data_concat.apply(
                    lambda x: str(x[New]) if str(x[New]) != str(x[Old]) else str(x[New]),
                    axis=1)

        elif colname.endswith("_y"):
            continue

        else:
            data_change[colname] = data_concat[colname]
            continue

    data_deleted["Flag"] = "Deleted"
    data_add["Flag"] = "New"
    data_change["Flag"] = "Updated"
    data_unchange["Flag"] = "No Change"
    # data_unchange["Flag"] = "Old"

    not_compare_key = list(map(lambda x: x.lower(), system_variables))
    not_compare_none = list(map(lambda x: x.lower(), not_compare_variables_list))

    if len(not_compare_dataframe_Old) > 0:
        data_deleted = data_deleted.merge(not_compare_dataframe_Old, how="left", on=not_compare_key)
    else:
        data_deleted = pd.concat([data_deleted, pd.DataFrame(columns=not_compare_none)], sort=False)

    data_other = pd.concat([data_add, data_change, data_unchange], ignore_index=True)

    if len(not_compare_dataframe_New) > 0:
        data_other = data_other.merge(not_compare_dataframe_New, how="left", on=not_compare_key)
    else:
        data_other = pd.concat([data_other, pd.DataFrame(columns=not_compare_none)], sort=False)

    if KEEP_DELETE_DATA:
        data_all = pd.concat([data_deleted, data_other], ignore_index=True)
    else:
        data_all = data_other
    return data_all

def rename_common(data: pd.DataFrame,seq:list,sas_label:dict):
    seq2 = list(map(lambda x: x.lower(), seq))
    keep_columns = [*seq2, *["Flag"]]
    data1 = data[keep_columns]
    SAS_LABEL = get_sas_label(sas_label)
    LABEL_LIST = dict(zip(data1.columns, SAS_LABEL))
    data_renamed = data1.rename(columns=LABEL_LIST)
    return data_renamed

def highlight_rows(x):
    if x["Flag"] == "deleted":
        return ['background-color: LightGray'] * len(x)
    elif x["Flag"] == "new":
        return ['background-color: MediumSeaGreen'] * len(x)
    else:
        return [None] * len(x)

def highlight_cells(x):
    if operator.contains(x, "->"):
        return 'background-color: yellow'
    else:
        return ''

def select_file(str_var):
    directory = filedialog.askdirectory()
    str_var.set(directory)

def get_compare_variables(data: pd.DataFrame):
    system =  ["project","environmentName","Subject","RecordPosition","StudyEnvSiteNumber", "projectid", "studyid", "subjectId", "StudySiteId", "siteid", "Site", "SiteNumber", "SiteGroup",
               "instanceId", "InstanceRepeatNumber", "folderid", "Folder", "FolderName", "FolderSeq", "TargetDays", "DataPageId",
               "PageRepeatNumber", "RecordDate", "RecordId", "MinCreated", "MaxUpdated", "SaveTS" ]
    compare_variables = []
    for column in data.columns:
        if not_in_list(column,system) and not column.endswith("_STD"):
            compare_variables.append(column)
    return compare_variables

def not_in_list(variable:str,variables:list):
    FLAG = True
    for i in variables:
        if i == variable:
            FLAG = False
            break
    return FLAG

def get_dataset_list(dataset_name):
    dataset_list = []
    for i in dataset_name:
        if i.endswith('.sas7bdat'):
            j = i.split('.sas7bdat')[0]
            dataset_list.append(j)
    return dataset_list

def get_dataset_list_excel(dataset_name):
    dataset_list = []
    for i in dataset_name:
        if i.endswith('.xlsx'):
            j = i.split('.xlsx')[0]
            dataset_list.append(j)
    return dataset_list

def get_sas_label(label_list:dict):
    new_list = []
    for key, item in label_list.items():
        new_list.append(item)

    def list_duplicates(seq):
        data = defaultdict(list)
        for i, item in enumerate(seq):
            data[item].append(i)
        return ((key, locs) for key, locs in data.items() if len(locs) > 1)

    def update_duplicate_data(seq, duplicates_data):
        result = seq
        for dup in duplicates_data:
            for i in range(len(dup[1])):
                result[dup[1][i]] = f"{dup[0]}_{i + 1}"
        return result

    duplicate_list = list_duplicates(new_list)
    result = update_duplicate_data(new_list, duplicate_list)
    return result

def get_label_list_fromALS(PATH_INPUT_eSpec) -> dict:
    content = pd.read_excel(PATH_INPUT_eSpec, sheet_name='eCRF')
    oid_list = content['变量'].tolist()
    oid_name = content['变量名'].tolist()
    oid_dict = dict(zip(oid_list, oid_name))
    return oid_dict

def filter_variables(path_input_variables, data: pd.DataFrame):
    filtered_list = [var for var in path_input_variables if var in data]
    return filtered_list

