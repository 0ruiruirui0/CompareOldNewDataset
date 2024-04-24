# -*- coding: utf-8 -*-
__author__ = "xiaoyu.li"

from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment, Protection

header_font = Font(name="Arial",
                   size=9,
                   bold=True,
                   italic=False,
                   vertAlign=None,
                   underline="none",
                   strike=False,
                   color="000000")

header_alignment = Alignment(horizontal="center",
                             vertical="center",
                             text_rotation=0,
                             wrap_text=True,
                             shrink_to_fit=False,
                             indent=0)

font = Font(name="Arial",
            size=9,
            bold=False,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="000000")
alignment_left = Alignment(horizontal="left",
                           vertical="center",
                           text_rotation=0,
                           wrap_text=False,
                           shrink_to_fit=False,
                           indent=0)

alignment_center = Alignment(horizontal="center",
                             vertical="center",
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)

fill_none = PatternFill(fill_type="solid",
                        start_color="FFFFFF",  # 前景色 - 白色
                        end_color="FFFFFF")  # 背景色 - 白色

fill_gray = PatternFill(fill_type="solid",  # 纯色填充
                        start_color="A9A9A9",  # 前景色 - 深灰色
                        end_color="A9A9A9")  # 背景色 - 深灰色

fill_yellow = PatternFill(fill_type="solid",  # 纯色填充
                          start_color="FFFF00",  # 前景色 - 纯黄色
                          end_color="FFFF00")  # 背景色 - 纯黄色

none_border = Side(border_style="thin",
                   color="BFBFBF")

border = Border(left=none_border,
                right=none_border,
                top=none_border,
                bottom=none_border,
                diagonal=none_border,
                diagonal_direction=none_border,
                outline=none_border,
                vertical=none_border,
                horizontal=none_border)

number_format = "General"

datetime_format = "yyyy mmm dd hh:mm:ss"

protection = Protection(locked=True,
                        hidden=False)


# named style for the header none
header_style_none = NamedStyle(name="header style none")
header_style_none.font = header_font
header_style_none.fill = fill_none
header_style_none.border = border
header_style_none.alignment = header_alignment
header_style_none.protection = protection

# named style for the header gray
header_style_gray = NamedStyle(name="header style gray")
header_style_gray.font = header_font
header_style_gray.fill = fill_gray
header_style_gray.border = border
header_style_gray.alignment = header_alignment
# header_style_gray.number_format = number_format
header_style_gray.protection = protection

value_style = NamedStyle(name="value style")
value_style.font = font
value_style.border = border
value_style.alignment = alignment_left
# value_style.number_format = number_format
value_style.protection = protection

# named style for the header flag
header_style_flag = NamedStyle(name="header style flag")
header_style_flag.font = header_font
header_style_flag.fill = fill_yellow
header_style_flag.border = border
header_style_flag.alignment = header_alignment
# header_style_gray.number_format = number_format
header_style_flag.protection = protection

# Content style
content_style = NamedStyle(name="content style")
content_style.font = Font(name="Arial",
                          size=18,
                          bold=True,
                          italic=False,
                          vertAlign=None,
                          underline="none",
                          strike=False,
                          color="000000")
content_style.alignment = alignment_center
content_style.border = border

content_heading_style = NamedStyle(name="content heading style")
content_heading_style.font = Font(name="Arial",
                                  size=12,
                                  bold=True,
                                  italic=False,
                                  vertAlign=None,
                                  underline="none",
                                  strike=False,
                                  color="000000")
content_heading_style.alignment = alignment_center
content_heading_style.border = border

content_value_style = NamedStyle(name="content value style")
content_value_style.font = Font(name="Arial",
                                size=12,
                                bold=False,
                                italic=False,
                                vertAlign=None,
                                underline="none",
                                strike=False,
                                color="000000")
content_value_style.alignment = alignment_left
content_value_style.border = border
